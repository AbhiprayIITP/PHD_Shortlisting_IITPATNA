import pandas as pd                           #Importing Required Libraries
import re
from datetime import date
import openpyxl


def WrtFile(row , index , sheet):                                                      #Function for writing selected students to xlsx file
    (sheet.cell(row,1)).value = row-1
    (sheet.cell(row,2)).value  = df_in['UserId'][index]
    (sheet.cell(row,3)).value  = df_in['159_e_full_name'][index]
    (sheet.cell(row,4)).value  = df_in['159_s_contact_address'][index]
    (sheet.cell(row,5)).value  = df_in['159_s_permanent_address'][index]
    (sheet.cell(row,6)).value  = df_in['159_r_phone_number'][index]
    (sheet.cell(row,7)).value  = df_in['159_d_email_id'][index]
    (sheet.cell(row,8)).value  = df_in['159_y_category'][index]
    (sheet.cell(row,9)).value  = df_in['159_h_date_of_birth'][index]
    (sheet.cell(row,10)).value = df_in['109_e_exam_score'][index]
    (sheet.cell(row,11)).value = df_in['109_o_valid_upto'][index]
    (sheet.cell(row,12)).value = df_in['109_k_exam_rank'][index]
    (sheet.cell(row,13)).value = df_in['159_d_1st_preference_for_area_of_research_for_phd'][index]
    (sheet.cell(row,14)).value = df_in['159_d_2nd_preference_for_area_of_research_for_phd'][index]
    (sheet.cell(row,15)).value = df_in['159_d_3rd_preference_for_area_of_research_for_phd'][index]
    (sheet.cell(row,16)).value = df_in['159_d_4th_preference_for_area_of_research_for_phd'][index]
    (sheet.cell(row,17)).value = df_in['102_e_percentage_of_marks_or_final_grade_point_average'][index]
    (sheet.cell(row,18)).value = df_in['102_g_year_of_passing'][index]
    (sheet.cell(row,19)).value = df_in['103_e_percentage_of_marks_or_final_grade_point_average'][index]
    (sheet.cell(row,20)).value = df_in['103_g_year_of_passing'][index]
    if(bool(btech.search(str(df_in['104_n_degree_or_examination'][index])))|bool(btech1.search(str(df_in['104_n_degree_or_examination'][index])))):
        (sheet.cell(row,21)).value = df_in['104_e_percentage_of_marks_or_final_grade_point_average'][index]
        (sheet.cell(row,22)).value = df_in['104_g_year_of_passing'][index]
    else:
        (sheet.cell(row,21)).value = ''
        (sheet.cell(row,22)).value = ''
    if(bool(bca.match(str(df_in['104_n_degree_or_examination'][index])))):
        (sheet.cell(row,25)).value = df_in['104_e_percentage_of_marks_or_final_grade_point_average'][index]
        (sheet.cell(row,26)).value = df_in['104_g_year_of_passing'][index]
    else:
        (sheet.cell(row,25)).value = ''
        (sheet.cell(row,26)).value = ''
    if(bool(mtech.search(str(df_in['214_n_degree_or_examination'][index])))|bool(mtech1.search(str(df_in['214_n_degree_or_examination'][index])))):
        (sheet.cell(row,23)).value = df_in['214_e_percentage_of_marks_or_final_grade_point_average'][index]
        (sheet.cell(row,24)).value = df_in['214_g_year_of_passing'][index]
    else:
        (sheet.cell(row,23)).value = ''
        (sheet.cell(row,24)).value = ''
    if(bool(mca.match(str(df_in['214_n_degree_or_examination'][index])))):
        (sheet.cell(row,27)).value = df_in['214_e_percentage_of_marks_or_final_grade_point_average'][index]
        (sheet.cell(row,28)).value = df_in['214_g_year_of_passing'][index]
    else:
        (sheet.cell(row,27)).value = ''
        (sheet.cell(row,28)).value = ''
    (sheet.cell(row,29)).value = df_in['216_o_refjournal_no'][index]
    #(sheet.cell(row,30)).value = df_in['159_y_category'][index]
    (sheet.cell(row,30)).value = ((str(df_in['106_d_position_held'][index])+  ' '  + str(df_in['106_n_name_of_organization'][index])) if ((str(df_in['106_d_position_held'][index])+  ' '  + str(df_in['106_n_name_of_organization'][index])) != 'nan nan') else '' )
    (sheet.cell(row,31)).value = ((str(df_in['161_d_position_held'][index])+  ' '  + str(df_in['161_n_name_of_organization'][index])) if ((str(df_in['161_d_position_held'][index])+  ' '  + str(df_in['161_n_name_of_organization'][index])) != 'nan nan') else '' )
    (sheet.cell(row,32)).value = ((str(df_in['162_d_position_held'][index])+  ' '  + str(df_in['162_n_name_of_organization'][index])) if ((str(df_in['162_d_position_held'][index])+  ' '  + str(df_in['162_n_name_of_organization'][index])) != 'nan nan') else '' )
    (sheet.cell(row,33)).value = ''    
    #(sheet.cell(row,34)).value = ''   
        
        
col = 1
row = 2
row_n = 2
wb = openpyxl.Workbook()  
wb_n =  openpyxl.Workbook()                                                      #Creatind Workbook Object
sheet = wb.active
sheet_n = wb_n.active
Headers = ['App. No.','ID','Name','Address 1','Address 2','Phone' ,'Email','Category','DoB'	,'Gate_score','Gate_score_validity','Gate_rank','Topics of interest 1','Topics of interest 2','Topics of interest 3','Topics of interest 4'	,'10 % marks or grade point average','10 Year of Passing','12 % marks or grade point average','12 Year of Passing','BTECH % marks or grade point average','BTECH Year of Passing','MTECH % marks or grade point average','MTECH Year of Passing','BCA % marks or grade point average','BCA Year of Passing','MCA % marks or grade point average','MCA Year of Passing','DD Number','exp 1','exp 2','exp 3','exp 4','remarks','FLAG']

for i in Headers:                                                              #Naming Of all Columns
    (sheet.cell(1,col)).value = i
    (sheet_n.cell(1,col)).value = i
    #c1.value = i
    col = col+1    
    
a = re.compile('computer' , re.I)                                              #Setting up various regex object to use for identification of key data
b = re.compile('cse' , re.I)
c = re.compile('information technology' , re.I)
d = re.compile('signal',re.I)
e = re.compile('cs',re.I)
f = re.compile('c.s.e',re.I)
g = re.compile('mathematics and computing',re.I)
q = re.compile('mathematics & computing',re.I)
h = re.compile('electronics',re.I)
j = re.compile('communication',re.I)
k = re.compile('digital',re.I)
l = re.compile('software' , re.I)
m = re.compile('IIT.*', re.I)
n  =re.compile('Indian Institute Of Technology.*',re.I)
o = re.compile('IISC' ,re.I)
p = re.compile('Indian Institute Of Science',re.I)
oo = re.compile('ISI',re.I)
pp = re.compile('Indian Statistical Institute',re.I)
mtech = re.compile('m.*tech',re.I|re.S)
mtech1 = re.compile('master.* of technology',re.I)
btech =  re.compile('b.*tech',re.I|re.S)
btech1 = re.compile('Bachelor Of Technology' ,re.I)
me = re.compile('m.?e',re.I)
msc = re.compile('m.sc',re.I)
mca = re.compile('mca',re.I)
ms = re.compile('m\.s',re.I)
ms1 = re.compile('ms',re.I)
bca = re.compile('bca',re.I)


input_file = 'raw_info_from_form_input_file.csv'                              
df_in = pd.read_csv(input_file,delimiter = ',')                               #Transferring data from csv file to pandas Datarame 

for ind in df_in.index:
    Gate_exemp = 0                                                               
    age_exemp = 0
    ugcpi,ugper,qcpi,qper,hscpi,hsper,experience = [0]*7
    Candidate = df_in['159_e_full_name'][ind]
    qfd = df_in['101_e_qualification_degree'][ind]
    qdis = df_in['101_e_discipline'][ind]
    IIT_check = df_in['101_y_name_and_place_of_institution_or_university'][ind]

    if(df_in['101_e_overall_percentage_of_marks_or_final_grade_point_average'][ind]<=10):
        qcpi = df_in['101_e_overall_percentage_of_marks_or_final_grade_point_average'][ind]
    elif(df_in['101_e_overall_percentage_of_marks_or_final_grade_point_average'][ind]>10):
        qper = df_in['101_e_overall_percentage_of_marks_or_final_grade_point_average'][ind]
    if(df_in['104_e_percentage_of_marks_or_final_grade_point_average'][ind]<=10):
        ugcpi = df_in['104_e_percentage_of_marks_or_final_grade_point_average'][ind]
    elif(df_in['104_e_percentage_of_marks_or_final_grade_point_average'][ind]>10):
        ugper = df_in['104_e_percentage_of_marks_or_final_grade_point_average'][ind]     
    imper =  df_in['103_e_percentage_of_marks_or_final_grade_point_average'][ind] 
    if(df_in['102_e_percentage_of_marks_or_final_grade_point_average'][ind]<=10):
        hscpi = df_in['102_e_percentage_of_marks_or_final_grade_point_average'][ind]
    elif(df_in['102_e_percentage_of_marks_or_final_grade_point_average'][ind]>10):
        hsper = df_in['102_e_percentage_of_marks_or_final_grade_point_average'][ind]      
    gate = df_in['215_n_have_you_written_the_gate_examination'][ind] 
    category = df_in['159_y_category'][ind] 
    gender =   df_in['159_r_gender'][ind] 
    age =   date.today().year - int(df_in['159_h_date_of_birth'][ind].strip()[-4:])
    
    if(df_in['109_e_exam_name'][ind] == 'GATE'):                                        #Checking if the exam given in GATE or UGC-NET
        gate_rank =  df_in['109_k_exam_rank'][ind]
    else:
        gate = 'No'
        
    pd = df_in['159_d_physically_handicapped'][ind]
    
    
    if(df_in['162_n_name_of_organization'][ind]!='' and type(df_in['162_n_name_of_organization'][ind]) is not float):                                                          #
                                                                                                                                                                               #                             
        if(df_in['162_k_end_date_of_work'][ind] == '' or type(df_in['162_k_end_date_of_work'][ind]) is float or df_in['162_k_end_date_of_work'][ind].strip()[-5] != '/'):      #
            experience =  date.today().year - int(df_in['162_k_start_date_of_work'][ind].strip()[-4:])                                                                         # 
        else:                                                                                                                                                                  #       
            experience =  int(df_in['162_k_end_date_of_work'][ind].strip()[-4:]) -  int(df_in['162_k_start_date_of_work'][ind].strip()[-4:])                                   # Calculating Experience from Start and End Date at current employer

    
    elif(df_in['161_n_name_of_organization'][ind]!='' and type(df_in['161_n_name_of_organization'][ind]) is not float):                                                        #
                                                                                                                                                                               #                             
        if(df_in['161_k_end_date_of_work'][ind] == '' or type(df_in['161_k_end_date_of_work'][ind]) is float or df_in['161_k_end_date_of_work'][ind].strip()[-5] != '/'):      #
            experience =  date.today().year - int(df_in['161_k_start_date_of_work'][ind].strip()[-4:])                                                                         # 
        else:                                                                                                                                                                  #       
            experience =  int(df_in['161_k_end_date_of_work'][ind].strip()[-4:]) -  int(df_in['161_k_start_date_of_work'][ind].strip()[-4:])                                   # 
                                                                                                                                                                               #          
                                                                                                                                                                               #              
    elif(df_in['106_n_name_of_organization'][ind]!='' and type(df_in['106_n_name_of_organization'][ind]) is not float):                                                        #          
                                                                                                                                                                               #           
        if(df_in['106_k_end_date_of_work'][ind] == '' or type(df_in['106_k_end_date_of_work'][ind]) is float or df_in['106_k_start_date_of_work'][ind].strip()[-5] != '/'):    #
            experience =  date.today().year - int(df_in['106_k_start_date_of_work'][ind].strip()[-4:])
        else:
            experience =  int(df_in['106_k_end_date_of_work'][ind].strip()[-4:]) -  int(df_in['106_k_start_date_of_work'][ind].strip()[-4:])
        

    
    
    if(a.search(qdis) == None and b.search(qdis) == None and c.search(qdis) == None and d.search(qdis) == None and e.match(qdis) == None and f.search(qdis) == None and g.search(qdis) == None and h.search(qdis) == None and j.search(qdis) == None and k.search(qdis) == None and l.search(qdis) == None and q.search(qdis) == None):
        disp_flag = 0          #Checking if correct discipline 
    else:
        disp_flag = 1
    
    
    if(m.match(IIT_check)!= None or n.search(IIT_check)!=None ):
       
        IIT_flag = 1          #Checking if IIT
    else:
        IIT_flag = 0
        
    if(o.match(IIT_check)!= None or p.search(IIT_check)!=None or oo.match(IIT_check)!=None or pp.match(IIT_check)!=None):
        IISC_flag = 1         #Checking if IISC or ISI
    else:
        IISC_flag = 0
        
    if(experience>=2):       #Removing Age criteria if experience>=2years
        age_exemp = 1
    
    
    Mtech_check = bool(mtech.search(qfd))|bool(mtech1.search(qfd))            #Various checks for identifying degree
    MS_check    = bool(ms.match(qfd))|bool(ms1.match(qfd))
    ME_check    = bool(me.search(qfd))
    Btech_check = bool(btech.search(qfd))|bool(btech1.search(qfd))
    MCA_check   = bool(mca.match(qfd))
    MSC_check   = bool(msc.search(qfd))
    
    
    if(Mtech_check or MS_check or ME_check):                                  #Relaxation Conditions as specified
        ugcpi_l,ugper_l,imper_l,hscpi_l,hsper_l = 6.5,60,60,6.5,60
        if(IIT_flag or IISC_flag or category == 'SC' or category == 'ST' or pd == 'Yes'):
            ugper_l = 55
            ugcpi_l = 6.0
        if(qcpi>=7.5 or qper>=70):                                            #If Cpi>7.5 or 70% allowing qualification of second class X and XII
            if(imper>=60):
                hscpi_l = 5.3
                hsper_l = 50
            elif(hscpi >= 6.5 or hsper >= 60):
                imper_l = 50
    
    
    
    if(Btech_check or MCA_check or MSC_check):
        ugcpi_l,ugper_l,imper_l,hscpi_l,hsper_l,gate_rank_l = 6.5,60,60,6.5,60,5000
        if(qcpi>=8.5 or qper>=80 or (Btech_check and IIT_flag and qcpi>=8.0)):   #If Cpi>8.5 or 80% allowing qualification of second class X and XII
            if(imper>=60):
                hscpi_l = 5.3
                hsper_l = 50
            elif(hscpi >= 6.5 or hsper >= 60):
                imper_l = 50
           
        if(Btech_check and IIT_flag and qcpi>=8.0):           
            Gate_exemp = 1
        
        if(category == 'SC' or category == 'ST' or pd == 'Yes'):
            gate_rank_l = 6000
         
        if((Btech_check or MCA_check or MSC_check) and experience>=2):          #Removing Gate qualification for experience>=2 years
            Gate_exemp = 1
             
    
    
    if((Mtech_check or MS_check or ME_check)  and (qcpi>=6.5 or qper>=60) and (ugcpi>=ugcpi_l or ugper>=ugper_l) and (imper>=imper_l) and (hscpi>= hscpi_l or hsper>=hsper_l) and (gate == 'Yes')):
        if((category=='General' and (age<=32 or age_exemp)) or ((category=='OBC Non Creamy Layer'  or category== 'EWS'  or gender == 'Female' or category == 'SC' or category == 'ST' or pd == 'Yes' ) and (age<=37 or age_exemp))):
            WrtFile(row,ind,sheet)
            row = row+1
            if(disp_flag == 0):
                (sheet.cell(row-1,35)).value = 'SHORTLISTED BUT PLEASE CHECK BRANCH ELIGIBILITY'
            if(experience>=2):
                (sheet.cell(row-1,34)).value = 'SPONSORED/PART-TIME'
                
        else:
            WrtFile(row_n,ind,sheet_n)
            row_n = row_n+1            
            
    elif((Btech_check or MCA_check or MSC_check)  and (qcpi>=8 or qper>=75)  and (ugcpi>=ugcpi_l or ugper>=ugper_l) and (imper>=imper_l) and (hscpi>= hscpi_l or hsper>=hsper_l) and (((gate=='Yes') and (gate_rank<=gate_rank_l)) or Gate_exemp)):
        if((category=='General' and (age<=28 or age_exemp)) or ((category=='OBC Non Creamy Layer'  or category== 'EWS'  or gender == 'Female' or category == 'SC' or category == 'ST' or pd == 'Yes' ) and (age<=33 or age_exemp))):
            WrtFile(row,ind,sheet)
            row = row+1
            if(disp_flag == 0):
                (sheet.cell(row-1,35)).value = 'SHORTLISTED BUT PLEASE CHECK BRANCH ELIGIBILITY'
            if(experience>=2):
                (sheet.cell(row-1,34)).value = 'SPONSORED/PART-TIME'
        else:
            WrtFile(row_n,ind,sheet_n)
            row_n = row_n+1  
            
    
    
    elif(Btech_check and (df_in['214_n_degree_or_examination'][ind] != '' and type(df_in['214_n_degree_or_examination'][ind]) is not float)):  # If qualification degree Btech , checking for Mtech passing criteria if possible
        if(IIT_flag or IISC_flag or category == 'SC' or category == 'ST' or pd == 'Yes'):                                                      
            ugper_l = 55
            ugcpi_l = 6.0
        if(qcpi>=7.5 or qper>=70):
            if(imper>=60):
                hscpi_l = 5.3
                hsper_l = 50
            elif(hscpi >= 6.5 or hsper >= 60):
                imper_l = 50
        if((qcpi>=6.5 or qper>=60) and (ugcpi>=ugcpi_l or ugper>=ugper_l) and (imper>=imper_l) and (hscpi>= hscpi_l or hsper>=hsper_l) and (gate == 'Yes')):
            if((category=='General' and (age<=32 or age_exemp)) or ((category=='OBC Non Creamy Layer'  or category== 'EWS'  or gender == 'Female' or category == 'SC' or category == 'ST' or pd == 'Yes' ) and (age<=37 or age_exemp))):
                WrtFile(row,ind,sheet)
                row = row+1
                #print(Candidate, 'New' ,df_in['UserId'][ind] )
                if(disp_flag == 0):
                    (sheet.cell(row-1,35)).value = 'SHORTLISTED BUT PLEASE CHECK BRANCH ELIGIBILITY'   
                if(experience>=2):
                    (sheet.cell(row-1,34)).value = 'SPONSORED/PART-TIME'
                    
            else:
                WrtFile(row_n,ind,sheet_n)
                row_n = row_n+1
        else:
            WrtFile(row_n,ind,sheet_n)
            row_n = row_n+1
            
                
    
    else:
        WrtFile(row_n,ind,sheet_n)
        row_n = row_n+1
        
    
            
            
            
wb.save('Final_Format.xlsx')                          #Saving to Excel         
wb_n.save('Not_Shortlisted_Final_Format.xlsx')            

        
        

        
        
        
    
    


      
    

   
    
    
    